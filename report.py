import openpyxl

Cell_name = input("Enter the cell name: ").lower()
# open the workbook
workbook = openpyxl.load_workbook('Cell 7 .xlsm') 

# select the Pressures sheet
worksheet = workbook['Pressures B'] 

# create an empty list to store row numbers
fail_rows = []

# loop through column K and look for the word "fail"
for row in range(1, worksheet.max_row + 1):
    cell_value = worksheet.cell(row=row, column=11).value
    if cell_value == "Fail":
        # add the row number to the list
        fail_rows.append(row)

new_numbers = []

for number in fail_rows:
    new_numbers.append("C" + str(number))


# print the list of row numbers
print("Transducers containing 'Fail' are located in Cells:", new_numbers)

for transducer in new_numbers:
    cell_value = worksheet[transducer].value
    print(f"{cell_value} has failed calibration.")

print(f"---In total, there are {len(new_numbers)} failed pressure transducers on {Cell_name}!---")
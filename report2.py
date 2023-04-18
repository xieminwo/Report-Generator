import openpyxl
import tkinter as tk 
from tkinter import filedialog 

root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename() # show an "Open" dialog box and return the path to the selected file 

print(file_path)

input("Press Enter to continue...")

Cell_name = input("Enter the cell name: ").lower()

# open the workbook
workbook = openpyxl.load_workbook(file_path)

# select the Pressures sheet
worksheet = workbook['Pressures A']

# create an empty list to store row numbers
fail_rows = []

# loop through column K and look for the word "fail"
for row in range(1, worksheet.max_row + 1):
    cell_value = worksheet.cell(row=row, column=11).value
    if cell_value == "Fail":
        # add the row number to the list
        fail_rows.append(row)

new_numbers = []

for number in fail_rows:
    new_numbers.append("C" + str(number))

failed_transducers = []

for transducer in new_numbers:
    cell_value = worksheet[transducer].value
    print(f"{cell_value} has failed calibration.")
    failed_transducers.append(cell_value)

print(
    f"---In total, there are {len(new_numbers)} failed pressure transducers on {Cell_name}!---")

print(failed_transducers)

# create a new sheet in the workbook
new_sheet = workbook.create_sheet(title="Failed Transducers")

# loop through the failed transducers and add them to cell A1 of the new sheet
for i, transducer in enumerate(failed_transducers):
    new_sheet.cell(row=1+i, column=1).value = transducer

# close the workbook
workbook.close()

# save the changes to the workbook
workbook.save('test.xlsm')
